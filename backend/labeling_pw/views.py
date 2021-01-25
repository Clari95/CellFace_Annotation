from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
import os, re, os.path
import numpy as np
import time
import openpyxl
from skimage import measure

from .code import createSegMask
from .code import createLabelMask
from .code import load_h5_WBC
from .code import addNewAreas
from .code import calculateAccuracy

# Create your views here.

def index(request):
    return HttpResponse("Hello, world. You're at the labeling PW index.")

def getDataLevel(request, user = '', tooltype = '', dataset_typ ='', level_num =''):
    if (level_num == 4 or level_num == 5):
        print('get Data for level 4 or 5')
        if level_num == 4:
            number = 1
            mounted = True
            image_section = 1
        else:
            number = 11
            mounted = False
            image_section = 2

        label_masks_level, image_series, highscore = getData(request, number, mounted, user, image_section, dataset_typ, tooltype, level_num)
        return JsonResponse({'label_masks': label_masks_level, 'images': image_series, 'highscore': highscore})

    if user == '':
        user = "testuser"
    # user = request.user.username
    SAVE_IMAGES = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'code/label_masks_temp')
    SEG_IMAGES = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'code/Seg_masks')
    #empty tmp folder for LabelMasks
    if level_num == 1:
        createLabelMask.deleteFiles(SAVE_IMAGES)
        while not os.path.exists(SEG_IMAGES + '_' + user + '_' + dataset_typ + '_' + tooltype):
            os.mkdir(SEG_IMAGES + '_' + user + '_' + dataset_typ+ '_' + tooltype)
            print("created folder for user")
        if os.path.exists(SEG_IMAGES + '_' + user + '_' + dataset_typ+ '_' + tooltype):
            print('delete existing folder')
            createLabelMask.deleteFiles(SEG_IMAGES + '_' + user + '_' + dataset_typ+ '_' + tooltype)

    SEG_Mask = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              'code/Tutorial/mask/'+ str(level_num) +'_Level')
    IMAGES = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            'code/Tutorial/img/'+ str(level_num) +'_Level')
    #results_img = np.load(SEG_IMAGES + '/' + 'ALL_SegMasks.npy')
    label_masks_level = np.empty((0, 384, 512), int)
    image_series = np.empty((0, 384, 512), int)
    k = 1
    if level_num != 3:
        num = 6
    else:
        num = 3
        # create tmp folder for labelregions.npy for adding regions

    while k < num:
        label_mask = np.load(SEG_Mask + '/' + str(k) + '_labelregions.npy')
        #save maks in tmp file
        np.save(SAVE_IMAGES + '/' + str(level_num) + 'level_' + str(k) + '_labelregions.npy', label_mask)
        label_mask = label_mask[np.newaxis, ...]
        label_masks_level = np.append(label_masks_level, np.atleast_3d(label_mask), axis=0)
        img = np.load(IMAGES + '/image_' + str(k) + '.npy')
        img = img[np.newaxis, ...]
        image_series = np.append(image_series, np.atleast_3d(img), axis=0)
        k = k + 1

    image_series = image_series.tolist()
    label_masks_level = label_masks_level.tolist()

    SAVED_Scores = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'code/Scores_' + tooltype + '.xlsx')
    wb = openpyxl.load_workbook(SAVED_Scores)
    ws = wb.get_sheet_by_name('Sheet1')
    row_count = ws.max_row
    for col in ws.iter_cols(min_row=3, min_col=1, max_row=row_count, max_col=2, values_only=True):
        col_points_overall = col
    highscore_overall = max(col_points_overall)

    cols = 9 + ((level_num - 1) * 12)
    for col_level in ws.iter_cols(min_row=3, min_col=cols, max_row=row_count, max_col=cols + 1, values_only=True):
        col_points = col_level
    highscore_level = max(col_points)
    highscore = [highscore_overall, highscore_level]
    #print('highscore: ' + str(highscore))

    return JsonResponse({'label_masks': label_masks_level, 'images': image_series, 'highscore': highscore})

def labels_received_level(request, data='', data_region ='', lastEle = '', number='', user='', dataset_typ='', level_num='', tooltype = ''):
    if request.is_ajax and request.method == "POST":
        print('level num: ', level_num)
        print('lastEle', lastEle)
        # get the form data
        #user = request.user.username
        #form = request.POST['dataset']
        #number = request.POST['image_number']
        #user = request.POST['user']
        if user == '':
            user = "testuser"
        print('label data of image: ', number, user)
        print('data: ', data, data_region)
        print(os.path.dirname(os.path.abspath(__file__)))
        folder = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              'code/label_masks_temp') #code/Tutorial/mask/first_SingelCells')
        SegMask, wrong_label, first_image = createSegMask.labelAreas(data, data_region, lastEle, number, user, dataset_typ, folder, level_num, tooltype)

        #folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'code/label_masks_temp')
        #SegMask, wrong_label, first_image = createSegMask.labelAreas(data, data_region, number, user, dataset_typ,
                                                                     #folder, level_num, tooltype)


        """accuracy:
            accuracy_areas: labled areas / all possible areas
            accuracy_pixel: labled pixel / all possibel pixel
            accuracy_wronglabel: correct labeled areas / all labled areas
        """
        #frist label in first image wrong: accuracy is zero
        if (wrong_label == True and first_image == True):
            accuracy = [0, 0, 0]
        else:
            accuracy_areas, accuracy_pixel, accurcay_correctlabel = calculateAccuracy.getAccuracy(SegMask, number, level_num, dataset_typ)
            #calculateAccuracy.dice_loss()
            accuracy = [accuracy_areas, accuracy_pixel, accurcay_correctlabel]
            print('accuracy: ', accuracy)
        #else:
         #   accuracy = [1, 1, 1]
        #SegMask = SegMask.tolist()

        """ save images from last two levels to compare it with solution"""
        if level_num == 4 or level_num == 5:
            SEG_IMAGES = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'code/Seg_masks' + '_' + user + '_' + dataset_typ + '_' + tooltype)
            results_img = np.empty((0, 384, 512), int)
            i = 1
            while i <= number:
                path = SEG_IMAGES + '/level'+str(level_num)+ '_' + str(i) +'_SegMask.npy'
                if os.path.exists(path):
                    #print('path: ', path)
                    n = np.load(path)
                    #diff = solution - n
                    #diff[]
                    n = n[np.newaxis, ...]
                    results_img = np.append(results_img, np.atleast_3d(n), axis=0)
                    # results_img.append(n)
                else:
                    m = np.zeros((1, 384, 512), int)
                    results_img = np.append(results_img, np.atleast_3d(m), axis=0)

                i = i + 1

            np.save(SEG_IMAGES + '/' + 'ALL_level_' + str(level_num) + '_SegMasks.npy', results_img)
            #print('length: ' + str(len(results_img)))



        #wrong_label = str(wrong_label)
        #print('wrongLabel: ' + str(wrong_label))

    return JsonResponse({'wrong_label': wrong_label, 'image_num': number, 'accuracy': accuracy})

def getData(request, number='', mounted='', user='', image_section='', dataset_typ ='', tooltype = '', level_num = ''):
    print('get data')
    if user == '':
        user = "testuser"
    #print('get data: ', number, level_num)
    #user = request.user.username
    SAVE_IMAGES = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'code/label_masks_temp')
    SEG_IMAGES = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'code/Seg_masks')

    #for adding test image adding areas in between
    SEG_Mask = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            'code/Tutorial/mask/3_Level')
    IMAGES = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          'code/Tutorial/img/3_Level')
    #TEST_IMAGES = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'code/test')
    #frist time moutned, create folder for user and gt highscore
    if mounted == 'true':
        createLabelMask.deleteFiles(SAVE_IMAGES)
        while not os.path.exists(SEG_IMAGES + '_' + user + '_' + dataset_typ + '_' + tooltype):
            os.mkdir(SEG_IMAGES + '_'+ user + '_' + dataset_typ + '_' + tooltype)
            print("created folder for user")
        if os.path.exists(SEG_IMAGES + '_'+ user + '_' + dataset_typ + '_' + tooltype) and level_num == 0:
            #print('no delete')
            createLabelMask.deleteFiles(SEG_IMAGES + '_'+ user + '_' + dataset_typ + '_' + tooltype)

        """get highscore"""
    SAVED_Scores = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'code/Scores_' + tooltype + '.xlsx')
    wb = openpyxl.load_workbook(SAVED_Scores)
    ws = wb.get_sheet_by_name('Sheet1')
    row_count = ws.max_row
    for col in ws.iter_cols(min_row=3, min_col=1, max_row=row_count, max_col=2, values_only=True):
        col_points_overall = col
    highscore_overall = max(col_points_overall)

    highscore = highscore_overall

    if level_num == 4 or level_num == 5:
        cols = 9 + ((level_num - 1) * 12)
        for col_level in ws.iter_cols(min_row=3, min_col=cols, max_row=row_count, max_col=cols + 1, values_only=True):
            col_points = col_level
        highscore_level = max(col_points)
        highscore = [highscore_overall, highscore_level]
    #print('highscore: ' + str(highscore))

    image_series, mean_image, celltyp = load_h5_WBC.prepareData(dataset_typ) #mean_image at the moment not used anymore
    label_masks_10 = np.empty((0, 384, 512), int)
    k = 0
    while k < 9:
        img = image_series[k + number - 1, :, :]
        mask, label_mask = createLabelMask.createMask(img, number + k, SAVE_IMAGES)
        label_mask = label_mask[np.newaxis, ...]
        label_masks_10 = np.append(label_masks_10, np.atleast_3d(label_mask), axis=0)
        k = k + 1

    print(image_section, level_num)
    if ((image_section != 1) or (level_num == 4)) and not(image_section == 2 and level_num == 5):
        print(' not change images')
        k = 9
        img = image_series[k + number - 1, :, :]
        mask, label_mask = createLabelMask.createMask(img, number + k, SAVE_IMAGES)
        label_mask = label_mask[np.newaxis, ...]
        label_masks_10 = np.append(label_masks_10, np.atleast_3d(label_mask), axis=0)
        img = image_series[int(number) - 1:int(number) + 9, :, :]
    else:
        print('change images')
        # change last images in section 0 with level 3 images 1 to check accuracy by adding areas.
        k = 1
        label_mask = np.load(SEG_Mask + '/' + str(k) + '_labelregions.npy')
        # save maks in tmp file as 20th image
        np.save(SAVE_IMAGES + '/20_labelregions.npy', label_mask)
        label_mask = label_mask[np.newaxis, ...]
        label_masks_10 = np.append(label_masks_10, np.atleast_3d(label_mask), axis=0)
        img = np.load(IMAGES + '/image_' + str(k) + '.npy')
        image_series = image_series[int(number) - 1:int(number) + 8, :, :]
        img = img[np.newaxis, ...]
        img = np.append(image_series, np.atleast_3d(img), axis=0)

      # 10 images
    image_10 = img.tolist()
    label_masks_10 = label_masks_10.tolist()
    print('send package of 10 images and LabelMask')
    if level_num == 4 or level_num == 5:
        return label_masks_10, image_10, highscore
    else:
        print('send JSON')
        return JsonResponse({'label_mask': label_masks_10, 'image': image_10, 'celltyp': celltyp, 'highscore': highscore}) #HttpResponse(response)#"Label Mask calculated and saved as .npy for image number: " + number)#, render(request, {"label_mask": label_mask}),

def labels_received(request, data='', data_region ='', lastEle = '', number='', user='', dataset_typ='', level_num='', tooltype=''):
    if request.is_ajax and request.method == "POST":
        print('in labels recieved')
        if user == '':
            user = "testuser"
        print('label data for image: ', number, user)
        print('data: ', data, data_region)
        folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'code/label_masks_temp')
        SegMask, wrong_label, first_image = createSegMask.labelAreas(data, data_region, lastEle, number, user, dataset_typ, folder, level_num, tooltype)
        #SegMask = SegMask.tolist()

        if (wrong_label == True & first_image == True) or level_num == 0:
            accuracy = [0, 0, 0]
        else:
            accuracy_areas, accuracy_pixel, accurcay_correctlabel = calculateAccuracy.getAccuracy(SegMask, number, level_num, dataset_typ)
            #calculateAccuracy.dice_loss()
            accuracy = [accuracy_areas, accuracy_pixel, accurcay_correctlabel]
            print('accuracy: ', accuracy)

        #labels = measure.label(SegMask)
        #print('number of labeled areas: ', labels.max())

        SEG_IMAGES = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'code/Seg_masks' + '_' + user + '_' + dataset_typ + '_' + tooltype)
        results_img = np.empty((0, 384, 512), int)
        i = 1
        while i <= number:
            path = SEG_IMAGES + '/level' +str(level_num) + '_' + str(i) +'_SegMask.npy'
            if os.path.exists(path):
                print('path: ', path)
                n = np.load(path)
                n = n[np.newaxis, ...]
                results_img = np.append(results_img, np.atleast_3d(n), axis=0)
                # results_img.append(n)
            else:
                m = np.zeros((1, 384, 512), int)
                results_img = np.append(results_img, np.atleast_3d(m), axis=0)

            i = i + 1

        if level_num == 4 or level_num == 5:
            np.save(SEG_IMAGES + '/' + 'ALL_level_' + str(level_num) + '_SegMasks.npy', results_img)
            #np.save(SEG_IMAGES + '/' + 'ALL_SegMasks.npy', results_img)
        else:
            np.save(SEG_IMAGES + '/' +'ALL_SegMasks.npy', results_img)

        results_img = results_img.tolist()

    return JsonResponse({'seg_masks': results_img, 'image_num': number, 'wrong_label': wrong_label, 'accuracy': accuracy}) #, 'celltyp': celltyp})
        #HttpResponse("label recieved und Segmentation Mask calculated for image number: " + str(number))

def new_areas(request, data='', number='', user='', dataset_typ ='', level_num =''):
    print('in new_areas')
    #data = request.POST['dataset']
    #number = request.POST['image_number']
    #user = request.POST['user']
    print("new data: " + data + dataset_typ + str(number) + str(level_num))
    if (level_num != 3 and number != 20):
        SAVE_IMAGES = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'code/label_masks_temp')
        image_series, mean_image, _ = load_h5_WBC.prepareData(dataset_typ)
        img = image_series[int(number) - 1, :, :]
        file = os.path.join(SAVE_IMAGES, str(number) + '_labelregions.npy')
    elif number == 20: #(level_num == 5):
        SAVE_IMAGES = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'code/label_masks_temp')
        IMAGES = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'code/Tutorial/img/3_Level')
        file = os.path.join(SAVE_IMAGES, str(number) + '_labelregions.npy')

        img = np.load(IMAGES + '/image_1.npy')
    else:
        SAVE_IMAGES = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'code/label_masks_temp')
        IMAGES = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'code/Tutorial/img/3_Level')
        file = os.path.join(SAVE_IMAGES, str(level_num) + 'level_' + str(number) + '_labelregions.npy')
        #k = 1
        #label_masks_level1 = np.empty((0, 384, 512), int)
        #image_series = np.empty((0, 384, 512), int)
        #label_mask = np.load(file)
        #label_mask = label_mask[np.newaxis, ...]
        #label_masks_level1 = np.append(label_masks_level1, np.atleast_3d(label_mask), axis=0)
        #img = np.load(IMAGES + '/image_' + str(k) + '.npy')
        # img = img[np.newaxis, ...]
        #img = img[np.newaxis, ...]
        #image_series = np.append(image_series, np.atleast_3d(img), axis=0)
        #print(image_series.shape)
        #print(img.shape)
        img = np.load(IMAGES +'/image_' + str(number) + '.npy')

    #print(image_series.shape)
    #print('image shape: ', img.shape)
    #img = np.squeeze(img, 0)
    #_, label_mask = createLabelMask.createMask(img, number, 'none')
    #print('image shape: ', type(img))
    #img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    print('newlabelmask of image ', number, SAVE_IMAGES) #% number % SAVE_IMAGES)
    LabelMask_edit, wrong_area = addNewAreas.start(img, number, data, SAVE_IMAGES)

    while not os.path.exists(file):
        time.sleep(1)
    if os.path.isfile(file):
        label_mask_before = np.load(file)   #load LabelMask
    else:
        raise ValueError("%s isn't a file!" % file)

    #print('File: ', file)
    newLabelMask = LabelMask_edit + label_mask_before
    #rewrite new LabelMask
    np.save(file, newLabelMask)
    #newLabelMask[newLabelMask != 0] = 1
    #print('new shape ', newLabelMask.shape)

    #addNewAreas.plot(newLabelMask, 'LabelMask', number, SAVE_IMAGES)
    #cv2.imwrite(SAVE_IMAGES + '/test.png', mask)
    #addNewAreas.plot(mask, 'LabelMask_mask', number, SAVE_IMAGES)

    newLabelMask = newLabelMask.tolist()

    return JsonResponse({'New_label_mask': newLabelMask, 'wrong_area': wrong_area})

def getDataReview(request, user='', dataset_typ='', tooltype ='', numberImages=''):
    print('get results of: ', user)
    user_back = user
    SEG_MASKS_user = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              'code/Seg_masks' + '_' + user + '_' + dataset_typ + '_' + tooltype)
    """getting seg maks results for showing in review"""
    if tooltype == 'basic':
        if os.path.isfile(SEG_MASKS_user + '/' + 'ALL_SegMasks.npy'):
            results_img = np.load(SEG_MASKS_user + '/' + 'ALL_SegMasks.npy')
        else:
            results_img = np.empty((0, 384, 512), int)

        data_ges = 0
        #only for game here zero
        leaders = [0, 0, 0]

    elif tooltype == 'game':
        if os.path.isfile(SEG_MASKS_user + '/' + 'ALL_level_4_SegMasks.npy'):
            reference_4 = np.load(SEG_MASKS_user + '/' + 'ALL_level_4_SegMasks.npy')
        else:
            reference_4 = np.empty((0, 384, 512), int)
        #reference_4 = np.load(SEG_IMAGES + '/' + 'ALL_level_4_SegMasks.npy')

        if os.path.isfile(SEG_MASKS_user + '/' + 'ALL_level_5_SegMasks.npy'):
            reference_5 = np.load(SEG_MASKS_user + '/' + 'ALL_level_5_SegMasks.npy')
            reference_5[0:reference_4.shape[0], :, :] = reference_4
        else:
            #reference_5 = np.empty((0, 384, 512), int)
            reference_5 = reference_4
        results_img = reference_5

        SAVED_Scores = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'code/Scores_' + tooltype + '.xlsx')
        wb = openpyxl.load_workbook(SAVED_Scores)
        ws = wb.get_sheet_by_name('Sheet1')
        row_count = ws.max_row
        data_ges = []
        for col in ws.iter_cols(min_row=row_count, min_col=0, max_row=row_count, max_col=71, values_only=True):
            #print(col)
            data_ges.append(col)
        #print('data ges: ' + str(data_ges))
        data_gesCOPY = data_ges # .copy()
        col_points_overall = []
        for col in ws.iter_cols(min_row=3, min_col=2, max_row=row_count, max_col=2, values_only=True):
            col_points_overallOrg = col
        print('points overall: ', col_points_overall)
        i=0
        index_max = []
        leaders = []
        col_points_overall = list(col_points_overallOrg)
        while i < 3:
            highscore_overall = max(col_points_overall)
            index = col_points_overallOrg.index(highscore_overall)
            #index_max.append(index)
            print(highscore_overall)
            print(col_points_overall)
            #if i < 2:
            col_points_overall.remove(highscore_overall)
            print(col_points_overall)
            i = i + 1
            leader = []
            for row in ws.iter_cols(min_row=index + 3, min_col=0, max_row=index + 3, max_col=71,
                                    values_only=True):
                leader.append(row)
            leaders.append(leader)
            #print(leaders)
        #print(leaders)
    else:
        print('tooltype false')

    #when in last images nothing is labled but the user has looked at it, it should be shown in review as empty
    #print('shape results_img:', results_img.shape[0])
    if results_img.shape[0] < numberImages - 1:
        m = np.zeros((1, 384, 512), int)
        while results_img.shape[0] < numberImages - 1:
            results_img = np.append(results_img, np.atleast_3d(m), axis=0)
    #print(results_img.shape[0])

    #print('AllSegmasks')
    #print(len(results_img))
    number_images = len(results_img)
    results_img = results_img.tolist()
    print('LEADERS', leaders)

    return JsonResponse({'SegMasksResults': results_img, 'numberOfimages': number_images, 'data_ges': data_gesCOPY, 'leader': leaders[0], 'leader2': leaders[1], 'leader3': leaders[2]})

def updateScores(request, tooltype='', user='', dataset_typ='', allScores = '', date =''):
    #calculate accuracy for 20 images
    print('UPDATESCORES')
    SEG_MASKS = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              'code/solution/Seg_masks_solution_Neutrophils')
    SEG_MASKS_user = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             'code/Seg_masks' + '_' + user + '_' + dataset_typ + '_' + tooltype)
    solution = np.load(SEG_MASKS + '/' + 'ALL_SegMasks.npy')
    allScores = allScores.split(",")
    print('updateScores in  excel')
    print(len(allScores))
    print(allScores)
    if len(allScores) > 24:
        allScores = allScores[:24]
        print('short:',  allScores)
    if os.path.isfile(SEG_MASKS_user + '/' + 'ALL_SegMasks.npy'):
        reference = np.load(SEG_MASKS_user + '/' + 'ALL_SegMasks.npy')
        print('size reference: ', reference.shape[0])
        #size = reference.shape[0]
    else:
        reference = np.empty((0, 384, 512), int)
    numberImages4 = int(allScores[0])
    #print(allScores)
    #print(numberImages4)
    print('number images: ', allScores[0], allScores[12])
    #if size < 9:
        #accuracy_4 = calculateAccuracy.getAccuracy_final(solution, reference[0:size, :, :], numberImages, 0)
    #else:
    #print('accuracy level 4')
    reference_copy = np.copy(reference)
    reference_copy2 = np.copy(reference)
    if int(allScores[12]) != 0:
        print('acc for level 4/ section 1')
        print(solution.shape[0])
        numberImages_ges = numberImages4 - 1 + int(allScores[12]) - 10
        accuracy_4 = calculateAccuracy.getAccuracy_final(solution, reference_copy[0:numberImages4 - 1, :, :], numberImages4 - 1, 0)

        #accuracy_4 = [0, 0, 0, 0]
    else:
        accuracy_4 = calculateAccuracy.getAccuracy_final(solution, reference[0:numberImages4, :, :], numberImages4, 0)
        numberImages_ges = numberImages4
        #accuracy_4 = [0, 0, 0, 0]

    if int(allScores[12]) != 0:
        print('acc for level 5/ section 2')
        print(solution.shape[0])
        numberImages5 = int(allScores[12]) - 10
        accuracy_5 = calculateAccuracy.getAccuracy_final(solution, reference_copy2[10:(numberImages5) + 10, :, :], numberImages5, 5)
        #accuracy_5 = [0, 0, 0, 0]
        print('accuracy level 4 and 5')
        #accuracy_all = calculateAccuracy.getAccuracy_final(solution, reference, (numberImages4 - 1) + numberImages5, 0)
        i = 0
        accuracy_all = [0, 0, 0, 0]
        while i < 4:
            accuracy_all[i] = (accuracy_4[i] * 10 + accuracy_5[i] * numberImages5) / (10 + numberImages5)
            #print(type(accuracy_4[i]), accuracy_5[i], numberImages5)
            print(((accuracy_4[i]) * 10 + (accuracy_5[i]) * numberImages5) / (10 + numberImages5))
            i = i + 1
        print('accuracy overall: ', accuracy_all)
    else:
        numberImages5 = 0
        accuracy_5 = [0,0,0,0]
        print('accuracy level 4 and 5')
        accuracy_all = accuracy_4
    #print('Accuracies: ', accuracy_4, accuracy_5, accuracy_all)
    SAVED_Scores = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'code/Scores_' + tooltype + '.xlsx')
    #missLabeldAreas = minusPoints / 5
    #newdata = [user, points, numberCells, numberRegions, addedAreas, missLabeldAreas, time, numberImages, accuracy_all[0], accuracy_all[1], accuracy_all[2], accuracy_all[3], dataset_typ,
    #           date]

    points_ges = 0
    time_ges = 0
    added_ges = 0
    number_incorrect = 0
    numberRegions_ges = 0
    numberCells_ges = 0
    #numberImages_ges = 0
    for i in [0, 12]:
        # print('points: ', float(allScores[i]))
        #numberImages_ges = numberImages_ges + float(allScores[i]) - 2
        #print(float(allScores[i]))
        numberCells_ges = numberCells_ges + float(allScores[i + 1])
        numberRegions_ges = numberRegions_ges + float(allScores[i + 2])
        number_incorrect = number_incorrect + float(allScores[i + 3])
        added_ges = added_ges + float(allScores[i + 4])
        time_ges = time_ges + float(allScores[i + 5])
        points_ges = points_ges + float(allScores[i + 6])

    newdata = [user, points_ges, dataset_typ]

    j = 0
    while (j < len(allScores)):
        newdata.append(float(allScores[j]))
        j = j + 1

    if newdata[15] != 0:
        newdata[15] = newdata[15] - 10
        newdata[3] = newdata[3] - 1

    data_ges = [numberImages_ges, numberCells_ges, numberRegions_ges, number_incorrect, added_ges, time_ges, points_ges]
    #print('data ges: ', data_ges)
    k = 0
    while (k < len(data_ges)):
        newdata.append(int(data_ges[k]))
        k = k + 1
    newdata.append(accuracy_all[0])
    newdata.append(accuracy_all[1])
    newdata.append(accuracy_all[2])
    newdata.append(accuracy_all[3])
    newdata.append(date)
    newdata[len(newdata) - 29:len(newdata) - 25] = accuracy_4
    newdata[len(newdata) - 17:len(newdata) - 13] = accuracy_5
    wb = openpyxl.load_workbook(SAVED_Scores)
    print('data written in excel: ', newdata)
    ws = wb.get_sheet_by_name('Sheet1')
    ws.append(newdata)
    wb.save(SAVED_Scores)
    wb.close()
    #accuracy = [accuracy_areas, accuracy_pixel, accurcay_wronglabeld]

    return JsonResponse({'Status': newdata, 'accuracy': accuracy_all})

def updateScoresGame(request, tooltype='', user='', dataset_typ='', allScores = '', date =''):
    #me/<str:tooltype>/<str:user>/<str:dataset_typ>/<str:allScores>'
    SEG_MASKS = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             'code/solution/Seg_masks_solution_Neutrophils')
    SEG_MASKS_user = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                  'code/Seg_masks' + '_' + user + '_' + dataset_typ + '_' + tooltype)
    solution = np.load(SEG_MASKS + '/' + 'ALL_SegMasks.npy')
    if os.path.isfile(SEG_MASKS_user + '/' + 'ALL_level_4_SegMasks.npy'):
        reference_4 = np.load(SEG_MASKS_user + '/' + 'ALL_level_4_SegMasks.npy')
    else:
        reference_4 = np.empty((0, 384, 512), int)

    if os.path.isfile(SEG_MASKS_user + '/' + 'ALL_level_5_SegMasks.npy'):
        reference = np.load(SEG_MASKS_user + '/' + 'ALL_level_5_SegMasks.npy')
        reference_5 = reference[10:19, :, :]
        reference[0:reference_4.shape[0], :, :] = reference_4
    else:
        reference = reference_4
        reference_5 = np.empty((0, 384, 512), int)


    print('updateScores in  excel')
    allScores = allScores.split(",")
    #print(len(allScores))
    print(allScores)

    print('number of images: ', allScores[36], allScores[48])

    print('size solution:', solution.shape[0])
    numberImages = int(allScores[48])
    accuracy_5_all = calculateAccuracy.getAccuracy_final(solution, reference_5, numberImages, 5)

    #need this to compare it with basic tool
    #reference[0:reference_4.shape[0], :, :] = reference_4
    #print(reference.shape[0])
    numberImages = int(allScores[36]) + (int(allScores[48]))  # both one image more
    accuracy_45_all = calculateAccuracy.getAccuracy_final(solution, reference, numberImages, 45)

    SAVED_Scores = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'code/Scores_' + tooltype + '.xlsx')
    #missLabeldAreas = minusPoints / 5
    points_ges = 0
    time_ges = 0
    added_ges = 0
    number_incorrect = 0
    numberRegions_ges = 0
    numberCells_ges = 0
    numberImages_ges  = 0
    for i in [0, 12, 24, 36, 48]:
        #print('points: ', float(allScores[i]))
        numberImages_ges = numberImages_ges + float(allScores[i])
        numberCells_ges = numberCells_ges + float(allScores[i + 1])
        numberRegions_ges = numberRegions_ges + float(allScores[i + 2])
        number_incorrect = number_incorrect + float(allScores[i + 3])
        added_ges = added_ges + float(allScores[i + 4])
        time_ges = time_ges + float(allScores[i + 5])
        points_ges = points_ges + float(allScores[i+6])

    print('points gesamt: ', points_ges)
    print('numbercells ges: ', numberCells_ges)

    newdata = [user, points_ges,  dataset_typ]

    j = 0
    while (j < len(allScores)):
        newdata.append(float(allScores[j]))
        j = j + 1

    newdata.append(date)

    data_ges = [numberImages_ges, numberCells_ges, numberRegions_ges, number_incorrect, added_ges, time_ges, points_ges]
    k = 0
    while (k < len(data_ges)):
        newdata.append(int(data_ges[k]))
        k = k + 1
    #print('newdata länge: ', len(newdata))
    print('INFO about acc 5 and if its written in data')
    print(accuracy_5_all)
    print(newdata)
    newdata[len(newdata)-13:len(newdata)-9] = accuracy_5_all
    print(newdata)
    wb = openpyxl.load_workbook(SAVED_Scores)
    print(newdata)
    ws = wb.get_sheet_by_name('Sheet1')
    ws.append(newdata)
    wb.save(SAVED_Scores)
    wb.close()
    print('data is written in excel sheet')

    return JsonResponse({'Status': newdata, 'accuracy5': accuracy_5_all, 'accuracy45': accuracy_45_all})

def getAccuracy(request, tooltype='', user='', dataset_typ='', level_num = '', numberImages=''):
    #translation from basic to game
    """
    if tooltype == 'basic':
        if level_num == 0:
            level_num = 4
        elif level_num == 1:
            level_num = 5
    """
    SEG_MASKS = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             'code/solution/Seg_masks_solution_Neutrophils')
    SEG_MASKS_user = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                  'code/Seg_masks' + '_' + user + '_' + dataset_typ + '_' + tooltype)
    solution = np.load(SEG_MASKS + '/' + 'ALL_SegMasks.npy')
    #noFile = False
    if os.path.isfile(SEG_MASKS_user + '/' + 'ALL_level_4_SegMasks.npy'):
        reference_4 = np.load(SEG_MASKS_user + '/' + 'ALL_level_4_SegMasks.npy')
    else:
        reference_4 = np.empty((0, 384, 512), int)
    if level_num == 4:
        reference = reference_4
    elif level_num == 5:
        if os.path.isfile(SEG_MASKS_user + '/' + 'ALL_level_5_SegMasks.npy'):
            reference_5 = np.load(SEG_MASKS_user + '/' + 'ALL_level_5_SegMasks.npy')
            print('SHAPE REF5: ', reference_5.shape[0])
            #reference = np.load(SEG_MASKS_user + '/' + 'ALL_level_5_SegMasks.npy')
            reference = reference_5[10:19, :, :]
            #reference = reference_5
        else:
            reference = np.empty((0, 384, 512), int)
    """else:
        if os.path.isfile(SEG_MASKS_user + '/' + 'ALL_level_5_SegMasks.npy'):
            reference_5 = np.load(SEG_MASKS_user + '/' + 'ALL_level_5_SegMasks.npy')
            reference = reference_5[0:reference_4.shape[0], :, :] = reference_4
        else:
            #reference = np.empty((0, 384, 512), int)
            reference = reference_4 """
    #print(reference.shape[0])
    # reference = np.append(reference_4, reference_5, axis=0)
    # print('shape ref: ', reference_5)
    accuracy_all = calculateAccuracy.getAccuracy_final(solution, reference, numberImages, level_num)

    print('sending: ', accuracy_all)
    return JsonResponse({'accuracy_all': accuracy_all})

def getAccuracy_level(request, level_num = '', numberImages='', acc_allImages=''):

    print('acc for images in level ', level_num, acc_allImages, numberImages)

    #if level_num == (4 or 5):
    #    getAccuracy(request, 'game', user, dataset_typ, level_num, numberImages)
    acc_allImages = acc_allImages.split(",")
    j=0
    acc_num = 0
    acc_pixel = 0
    acc_correct = 0
    while j < len(acc_allImages):
        acc_num = acc_num + float(acc_allImages[j])
        acc_pixel = acc_pixel + float(acc_allImages[j+1])
        acc_correct = acc_correct + float(acc_allImages[j+2])
        j = j + 3

    acc_num = acc_num / numberImages
    acc_pixel = acc_pixel / numberImages
    acc_correct = acc_correct / numberImages

    average = (acc_num + acc_correct) / 2
    accuracy_all = [acc_num, acc_pixel, acc_correct, average]
    #print(acc_num, acc_pixel, acc_correct, average)
    print('accureacy all: ', accuracy_all)

    return JsonResponse({'accuracy_all': accuracy_all})
